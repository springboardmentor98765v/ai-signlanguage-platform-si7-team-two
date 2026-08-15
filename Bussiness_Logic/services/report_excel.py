"""
Report Excel Generator — Milestone 4, Day 4.

Generates Excel reports for:
1. Learning
2. Assessment
3. Accuracy
4. Certification

Uses openpyxl, the same library already used by
the existing Progress Report Excel generator.
"""

from pathlib import Path
from openpyxl import Workbook


REPORTS_DIR = Path("generated_reports")
REPORTS_DIR.mkdir(exist_ok=True)


# ---------------------------------------------------------
# 1. LEARNING REPORT EXCEL
# ---------------------------------------------------------

def generate_learning_report_excel(
    learner_name: str,
    report_data: dict,
    user_id: str,
) -> str:

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Learning Report"

    sheet.append(["Learner Learning Report"])

    sheet.append([])
    sheet.append(["Learner Name", learner_name])
    sheet.append(["User ID", user_id])

    sheet.append([])

    sheet.append([
        "Total Sessions",
        report_data["total_sessions"],
    ])

    sheet.append([
        "Completed Sessions",
        report_data["completed_sessions"],
    ])

    sheet.append([
        "Total Practice Time (Seconds)",
        report_data["total_practice_time"],
    ])

    sheet.append([
        "Total Assessments",
        report_data["total_assessments"],
    ])

    sheet.append([
        "Letters Attempted",
        ", ".join(
            report_data["attempted_letters"]
        ) or "None",
    ])

    file_path = (
        REPORTS_DIR
        / f"Learning_Report_{user_id}.xlsx"
    )

    workbook.save(file_path)

    return str(file_path)


# ---------------------------------------------------------
# 2. ASSESSMENT REPORT EXCEL
# ---------------------------------------------------------

def generate_assessment_report_excel(
    learner_name: str,
    report_data: dict,
    user_id: str,
) -> str:

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Assessment Report"

    sheet.append(["Assessment Report"])

    sheet.append([])
    sheet.append(["Learner Name", learner_name])
    sheet.append(["User ID", user_id])

    sheet.append([])

    sheet.append([
        "Total Assessments",
        report_data["total_assessments"],
    ])

    sheet.append([
        "Average Score (%)",
        report_data["average_score"],
    ])

    sheet.append([
        "Correct Assessments",
        report_data["correct_assessments"],
    ])

    sheet.append([
        "Incorrect Assessments",
        report_data["incorrect_assessments"],
    ])

    sheet.append([])

    sheet.append([
        "Expected Sign",
        "Predicted Sign",
        "Score (%)",
        "Correct",
    ])

    for assessment in report_data[
        "assessment_scores"
    ]:

        sheet.append([
            assessment["expected_sign"],
            assessment["predicted_sign"],
            assessment["score"],
            assessment["is_correct"],
        ])

    file_path = (
        REPORTS_DIR
        / f"Assessment_Report_{user_id}.xlsx"
    )

    workbook.save(file_path)

    return str(file_path)


# ---------------------------------------------------------
# 3. ACCURACY REPORT EXCEL
# ---------------------------------------------------------

def generate_accuracy_report_excel(
    learner_name: str,
    report_data: dict,
    user_id: str,
) -> str:

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Accuracy Report"

    sheet.append(["Accuracy Report"])

    sheet.append([])
    sheet.append(["Learner Name", learner_name])
    sheet.append(["User ID", user_id])

    sheet.append([])

    sheet.append([
        "Overall Accuracy (%)",
        report_data["overall_accuracy"],
    ])

    sheet.append([])

    sheet.append([
        "Letter",
        "Accuracy (%)",
    ])

    for letter, score in report_data[
        "letter_accuracy"
    ].items():

        sheet.append([
            letter,
            score,
        ])

    sheet.append([])

    sheet.append([
        "Strong Letters",
        ", ".join(
            report_data["strong_letters"]
        ) or "None",
    ])

    sheet.append([
        "Weak Letters",
        ", ".join(
            report_data["weak_letters"]
        ) or "None",
    ])

    file_path = (
        REPORTS_DIR
        / f"Accuracy_Report_{user_id}.xlsx"
    )

    workbook.save(file_path)

    return str(file_path)


# ---------------------------------------------------------
# 4. CERTIFICATION REPORT EXCEL
# ---------------------------------------------------------

def generate_certification_report_excel(
    learner_name: str,
    report_data: dict,
    user_id: str,
) -> str:

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Certification Report"

    sheet.append(["Certification Report"])

    sheet.append([])
    sheet.append(["Learner Name", learner_name])
    sheet.append(["User ID", user_id])

    sheet.append([])

    sheet.append([
        "Certification Status",
        report_data["certificate_status"],
    ])

    sheet.append([
        "Certificates Earned",
        report_data["certificates_earned"],
    ])

    sheet.append([])

    sheet.append([
        "Certificate Code",
        "Average Score (%)",
        "Lessons Completed",
        "Issued At",
        "Valid",
    ])

    for certificate in report_data[
        "certificates"
    ]:

        sheet.append([
            certificate["certificate_code"],
            certificate["average_score"],
            certificate["lessons_completed"],
            str(certificate["issued_at"])
            if certificate["issued_at"]
            else "-",
            certificate["is_valid"],
        ])

    file_path = (
        REPORTS_DIR
        / f"Certification_Report_{user_id}.xlsx"
    )

    workbook.save(file_path)

    return str(file_path)